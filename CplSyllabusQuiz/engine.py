import openpyxl as xl
from pathlib import Path
import random
class CsqEngine(object):
     def __init__(self):
          self.isChosen=[None]*105
          self.isReport=[None]*105
          self.isExpected=[None]*105
          self.grade=[None]*105
          self.eMark=[None]*105
          self.letters=["A","B","C","D"]
          self.bible=["Romans 12:2 Do not be conformed to this world, but be transformed by the renewing of your mind, so that you may prove what is the good, well-pleasing, and perfect will of God."
          ,"Matthew 24:36-37 But no one knows of that day and hour, not even the angels of heaven, but my Father only. As the days of Noah were, so will be the coming of the Son of Man."
          ,"Psalm 27:1 The Lord is my light and my salvation. Whom shall I fear?The Lord is the strength of my life. Of whom shall I be afraid?"
          ,"Hebrews 13:2 Do not forget to show hospitality to strangers, for in doing so, some have entertained angels without knowing it."
          ,"1 Chronicles 16:11 Seek the Lord and his strength. Seek his face forever more."
          ,"Romans 15:13 May the God of hope fill you with all joy and peace in believing, so that by the power of the Holy Spirit you may abound in hope."]

          self.coursesList=[]
          self.detailList=[]
          self.studiesList=[]

          self.courseBank=Path('data','courseBank.xlsx')
          self.studyBank=Path('data','studyBank.xlsx')

          self.courses=xl.load_workbook(self.courseBank)
          self.studyList=xl.load_workbook(self.studyBank)

          self.course=self.courses.active
          self.studies=self.studyList.active
          
          self.coursedLoaded=False
          self.studiesLoaded=False

     def getStudyList(self):
          try:
               if self.studiesLoaded == False:
                    for column in self.studies.iter_cols(min_col=1,max_col=1):
                         for cell in column:
                              self.studiesList.append(cell.value)
                    self.studiesLoaded=True
               
          except Exception as e:
               print(e)
          return self.studiesList


     def getCourse(self):
          try:
               if self.coursedLoaded == False:
                    for column in self.course.iter_cols(min_col=1,max_col=1):
                         for cell in column:
                              self.coursesList.append(cell.value)
                    self.coursedLoaded=True
             
          except Exception as e:
               print(e)
          return self.coursesList

     def getCourseDetail(self,x):
          try:
               for column in self.course.iter_cols(min_col=2,max_col=2):
                    for cell in column:
                         self.detailList.append(cell.value)
          except Exception as e:
               print(e)
          return self.detailList[x]
     def getRandomQuote(self):
          self.verse=random.randint(0,5)
          return self.bible[self.verse]
     def getQuestionBank(self,loc,xlFile,Row,x):

          
          questionBank=Path(loc,xlFile)
          questions=xl.load_workbook(questionBank)
          question=questions.active
          questionList=[]
          questionAndDetail=[]
          answerIndex=[]
          try:
               for column in question.iter_cols(min_col=2,max_col=2,min_row=Row,max_row=Row):
                    for cell in column:
                         answerIndex.append(cell.value)
               for column in question.iter_cols(min_col=1,max_col=1,min_row=Row,max_row=Row):
                    for cell in column:
                         questionAndDetail.append(cell.value)
               for sections in questionAndDetail[0].split('|'):
                    questionList.append(sections)
          except Exception as e:
               print(e)
          if x==-2: #Get scenario
               if questionList[0] == "isCourseEnd":
                    return -1
               else:
                    return questionList[0]
          elif x==-1: #Get question
               return questionList[1]
          elif x==0: #Get A
               return questionList[2]
          elif x==1: #Get B
               return questionList[3]
          elif x==2: #Get C
               return questionList[4]
          elif x==3: #Get D
               return questionList[5]
          elif x==4: #Get Correct
               return int(answerIndex[0])
     
     def getAnswers(self,chosen,expected,x,started):
          try:
               if chosen != -1 and expected != -1 and started != -1:
                    if chosen!=expected:
                         self.eMark[started]="❌"
                    else:
                         self.eMark[started]="✅"
                         self.grade[started]=1
                    self.isChosen[started]=self.letters[chosen]
                    self.isExpected[started]=self.letters[expected]
                    
                    self.isReport[started]=self.letters[chosen]+"\t\t\t"+self.letters[expected]+"\t"+self.eMark[started]
          except Exception as e:
               print(e)
          if x == 1:
               return self.isReport
     def getGrade(self,x):
          A=0
          for _all in self.grade:
               if not _all:
                    continue
               else:
                    A=A+int(_all)
          X=float(x)
          Y=float(A)
          Z=(Y/X)*100
          return round(Z,3)
     def reset(self):
          self.isChosen=[None]*105
          self.isReport=[None]*105
          self.isExpected=[None]*105
          self.eMark=[None]*105
          self.grade=[None]*105
          
